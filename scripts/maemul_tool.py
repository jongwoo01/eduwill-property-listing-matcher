#!/usr/bin/env python3
"""Deterministic profile, CSV/XLSX search, and safe mutation tool for eduwill-property-listing-matcher."""

from __future__ import annotations

import argparse
import csv
import fcntl
import hashlib
import io
import json
import math
import os
import re
import shutil
import stat
import sys
import tempfile
from contextlib import contextmanager
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils.cell import get_column_letter, range_boundaries
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover - reported at the command boundary
    Workbook = None
    load_workbook = None
    Font = None
    PatternFill = None
    get_column_letter = None
    range_boundaries = None
    DataValidation = None


LISTING_COLUMNS = [
    "번호",
    "상태",
    "종류",
    "거래",
    "지역",
    "동네",
    "단지명",
    "동호",
    "매매가(만원)",
    "보증금(만원)",
    "월세(만원)",
    "관리비(만원)",
    "전용(㎡)",
    "방수",
    "욕실",
    "층",
    "총층",
    "향",
    "준공",
    "주차(대)",
    "반려",
    "옵션",
    "입주가능",
    "인근역",
    "역도보(분)",
    "초등학교",
    "초등도보(분)",
    "접수일",
]

DETAIL_COLUMNS = [
    "번호",
    "소유자",
    "연락처",
    "권리관계",
    "임대차현황",
    "용도지역",
    "건폐율·용적률",
    "위반건축물",
    "주차시설유형",
    "관리상세",
    "시설상태",
    "옵션상세",
    "학군상세",
    "비선호시설",
    "공시가격",
    "취득조세",
    "특약·메모",
    "접수경로",
    "계약일",
    "실제계약금액(만원)",
]

NUMERIC_FIELDS = {
    "매매가(만원)",
    "보증금(만원)",
    "월세(만원)",
    "관리비(만원)",
    "전용(㎡)",
    "방수",
    "욕실",
    "층",
    "총층",
    "준공",
    "주차(대)",
    "역도보(분)",
    "초등도보(분)",
    "실제계약금액(만원)",
}

NONNEGATIVE_NUMERIC_FIELDS = NUMERIC_FIELDS - {"층"}
FORMULA_PREFIXES = ("=", "+", "-", "@")

STATUS_VALUES = {"진행", "보류", "완료"}
TRANSACTION_VALUES = {"매매", "전세", "월세", "?"}
YN_UNKNOWN_VALUES = {"Y", "N", "?"}
UNKNOWN_VALUES = {"", "?"}
SUPPORTED_OPERATORS = {"eq", "ne", "in", "not-in", "contains", "lte", "gte", "between"}
# 텍스트 필드에 크기 비교를 걸면 사전순으로 계산돼 조용히 무의미한 결과가 나온다.
ORDERED_ONLY_OPERATORS = {"lte", "gte", "between"}

DETAIL_SOURCE_ALIASES = {
    "공시가격(만원)": "공시가격",
}

LEGACY_ONLY_FIELDS = {
    "학군(분)",
    "접수일",
}


class ToolError(Exception):
    pass


def emit(payload: dict[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, indent=2))


def fail(message: str, **details: Any) -> None:
    raise ToolError(json.dumps({"message": message, **details}, ensure_ascii=False))


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def resolve_path(raw: str) -> Path:
    return Path(raw).expanduser().resolve()


def require_openpyxl() -> None:
    if load_workbook is None or Workbook is None:
        fail("로컬 Excel 기능에는 openpyxl이 필요합니다", hint="python3 -m pip install openpyxl")


def require_xlsx_path(path: Path) -> None:
    if path.suffix.casefold() != ".xlsx":
        fail("로컬 매물장은 .xlsx 파일이어야 합니다", path=str(path))


def default_profile_store() -> Path:
    codex_root = Path(os.environ.get("CODEX_HOME", Path.home() / ".codex")).expanduser()
    canonical = (codex_root / "eduwill-property-listing-matcher" / "profiles.json").resolve()
    legacy_candidates = (
        (codex_root / "edwill-property-listing-matcher" / "profiles.json").resolve(),
        (codex_root / "maemul-matching" / "profiles.json").resolve(),
    )
    if not canonical.exists():
        for legacy in legacy_candidates:
            if legacy.is_file():
                write_profile_store(canonical, load_profile_store(legacy))
                break
    return canonical


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def fsync_directory(path: Path) -> None:
    descriptor = os.open(path, os.O_RDONLY)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def parse_json(raw: str, label: str) -> Any:
    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        fail(f"{label} JSON을 해석할 수 없습니다", error=str(exc))


def is_unknown(value: Any) -> bool:
    return value is None or str(value).strip() in UNKNOWN_VALUES


def parse_number(value: Any, field: str) -> float:
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError):
        fail("숫자 필드 값이 올바르지 않습니다", field=field, value=value)
    if not math.isfinite(number):
        fail("숫자 필드는 유한한 값이어야 합니다", field=field, value=value)
    if field in NONNEGATIVE_NUMERIC_FIELDS and number < 0:
        fail("숫자 필드는 음수일 수 없습니다", field=field, value=value)
    if field == "층" and not -20 <= number <= 200:
        fail("층 값이 허용 범위를 벗어났습니다", field=field, value=value)
    if field == "준공" and not 1800 <= number <= date.today().year + 10:
        fail("준공 연도가 허용 범위를 벗어났습니다", field=field, value=value)
    return number


def normalize_input_value(field: str, value: Any) -> str:
    """Normalize JSON scalars and neutralize spreadsheet formulas in text fields."""
    if value is None:
        return "?"
    if isinstance(value, (dict, list)):
        fail("필드 값은 객체나 배열일 수 없습니다", field=field, value=value)
    raw = str(value)
    if field not in NUMERIC_FIELDS and raw.lstrip().startswith(FORMULA_PREFIXES):
        return "'" + raw
    return raw


def reject_unsafe_formula(value: str, field: str, row_number: int) -> None:
    if field in NUMERIC_FIELDS or is_unknown(value):
        return
    if value.lstrip().startswith(FORMULA_PREFIXES):
        fail(
            "스프레드시트 수식으로 실행될 수 있는 텍스트입니다",
            field=field,
            row=row_number,
            value=value,
            hint="앞에 작은따옴표(')를 붙여 일반 텍스트로 저장하세요",
        )


def validate_iso_date(value: str, field: str, row_number: int) -> None:
    if value in UNKNOWN_VALUES:
        return
    try:
        date.fromisoformat(value)
    except ValueError:
        fail("날짜는 YYYY-MM-DD 또는 ? 형식이어야 합니다", field=field, row=row_number, value=value)


def validate_rows(rows: list[dict[str, str]], kind: str) -> dict[str, Any]:
    columns = LISTING_COLUMNS if kind == "listing" else DETAIL_COLUMNS
    seen: set[str] = set()
    warnings: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=2):
        if set(row) != set(columns):
            fail("행 컬럼이 스키마와 다릅니다", row=index)
        item_id = row["번호"].strip()
        if not item_id or item_id == "?":
            fail("번호는 비어 있거나 ?일 수 없습니다", row=index)
        if item_id in seen:
            fail("중복 번호가 있습니다", row=index, item_id=item_id)
        seen.add(item_id)

        for field in columns:
            reject_unsafe_formula(row[field], field, index)

        for field in NUMERIC_FIELDS.intersection(columns):
            if not is_unknown(row[field]):
                parse_number(row[field], field)

        if kind == "listing":
            if row["상태"] not in STATUS_VALUES:
                fail("상태 값이 올바르지 않습니다", row=index, value=row["상태"])
            if row["거래"] not in TRANSACTION_VALUES:
                fail("거래 값이 올바르지 않습니다", row=index, value=row["거래"], allowed=sorted(TRANSACTION_VALUES))
            for field in ("반려", "옵션"):
                if row[field] not in YN_UNKNOWN_VALUES:
                    fail("Y/N/? 필드 값이 올바르지 않습니다", row=index, field=field, value=row[field])
            validate_iso_date(row["접수일"].strip(), "접수일", index)
            if row["입주가능"] not in {"", "?", "즉시", "협의"}:
                validate_iso_date(row["입주가능"].strip(), "입주가능", index)
            for required in ("종류", "거래", "지역"):
                if is_unknown(row[required]):
                    warnings.append({"row": index, "field": required, "warning": "검색 핵심값 미확인"})
            if any(is_unknown(row[field]) for field in ("종류", "거래", "지역")) and row["상태"] != "보류":
                fail("검색 핵심값이 미확인인 행은 상태=보류여야 합니다", row=index, item_id=item_id)
        else:
            validate_iso_date(row["계약일"].strip(), "계약일", index)

    return {"rows": len(rows), "warnings": warnings}


def read_csv_table(path: Path, kind: str) -> tuple[list[dict[str, str]], dict[str, Any]]:
    rows, summary, _ = read_csv_snapshot(path, kind)
    return rows, summary


def read_csv_snapshot(path: Path, kind: str) -> tuple[list[dict[str, str]], dict[str, Any], str]:
    """Parse and hash one immutable byte snapshot of a CSV file."""
    if not path.is_file():
        fail("CSV 파일을 찾을 수 없습니다", path=str(path))
    expected = LISTING_COLUMNS if kind == "listing" else DETAIL_COLUMNS
    try:
        content = path.read_bytes()
        text = content.decode("utf-8-sig")
        with io.StringIO(text, newline="") as handle:
            reader = csv.DictReader(handle)
            actual = reader.fieldnames or []
            if actual != expected:
                fail("CSV 헤더가 표준 스키마와 다릅니다", path=str(path), expected=expected, actual=actual)
            rows = [dict(row) for row in reader]
    except UnicodeDecodeError as exc:
        fail("CSV가 UTF-8로 읽히지 않습니다", path=str(path), error=str(exc))
    summary = validate_rows(rows, kind)
    return rows, summary, hashlib.sha256(content).hexdigest()


def workbook_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time().isoformat() == "00:00:00" else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def read_workbook_snapshot(
    path: Path,
    kind: str,
    sheet_name: str,
) -> tuple[list[dict[str, str]], dict[str, Any], str]:
    require_openpyxl()
    require_xlsx_path(path)
    if not path.is_file():
        fail("Excel 매물장을 찾을 수 없습니다", path=str(path))
    snapshot_sha = sha256_file(path)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            fail("Excel 매물장에 필요한 탭이 없습니다", path=str(path), sheet=sheet_name)
        sheet = workbook[sheet_name]
        expected = LISTING_COLUMNS if kind == "listing" else DETAIL_COLUMNS
        actual = [
            workbook_cell_text(cell.value)
            for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=len(expected)))
        ]
        if actual != expected:
            fail("Excel 탭 헤더가 표준 스키마와 다릅니다", path=str(path), sheet=sheet_name, expected=expected, actual=actual)
        rows: list[dict[str, str]] = []
        for values in sheet.iter_rows(min_row=2, max_col=len(expected), values_only=True):
            if all(value is None or str(value).strip() == "" for value in values):
                continue
            rows.append({field: workbook_cell_text(value) for field, value in zip(expected, values)})
    finally:
        workbook.close()
    summary = validate_rows(rows, kind)
    return rows, summary, snapshot_sha


def validate_workbook_pair(path: Path, listing_sheet: str, detail_sheet: str) -> None:
    listing_rows, _, listing_sha = read_workbook_snapshot(path, "listing", listing_sheet)
    detail_rows, _, detail_sha = read_workbook_snapshot(path, "detail", detail_sheet)
    if listing_sha != detail_sha:
        fail("Excel 매물장이 검증 중 변경되었습니다", path=str(path))
    listing_ids = {row["번호"] for row in listing_rows}
    orphan_ids = sorted(row["번호"] for row in detail_rows if row["번호"] not in listing_ids)
    if orphan_ids:
        fail("매물 탭에 없는 번호의 상세행이 있습니다", path=str(path), item_ids=orphan_ids[:20])


def style_workbook_sheet(sheet: Any, columns: list[str]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(columns)).coordinate}"
    fill = PatternFill("solid", fgColor="1F4E78")
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(1, index, column)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        sheet.column_dimensions[cell.column_letter].width = min(max(len(column) + 4, 12), 24)


def command_init_workbook(args: argparse.Namespace) -> dict[str, Any]:
    require_openpyxl()
    path = resolve_path(args.workbook)
    require_xlsx_path(path)
    if args.listing_sheet == args.detail_sheet:
        fail("검색용 탭과 상세 탭 이름은 달라야 합니다")
    if path.exists():
        fail("기존 Excel 파일을 덮어쓰지 않습니다", path=str(path))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    listing = workbook.active
    listing.title = args.listing_sheet
    detail = workbook.create_sheet(args.detail_sheet)
    style_workbook_sheet(listing, LISTING_COLUMNS)
    style_workbook_sheet(detail, DETAIL_COLUMNS)
    status_validation = DataValidation(type="list", formula1='"진행,보류,완료"', allow_blank=False)
    yn_validation = DataValidation(type="list", formula1='"Y,N,?"', allow_blank=False)
    listing.add_data_validation(status_validation)
    listing.add_data_validation(yn_validation)
    status_validation.add("B2:B1048576")
    yn_validation.add("U2:V1048576")

    descriptor, raw_temporary = tempfile.mkstemp(prefix=f".{path.stem}.", suffix=".xlsx", dir=path.parent)
    os.close(descriptor)
    temporary = Path(raw_temporary)
    published = False
    try:
        workbook.save(temporary)
        workbook.close()
        os.chmod(temporary, 0o600)
        read_workbook_snapshot(temporary, "listing", args.listing_sheet)
        read_workbook_snapshot(temporary, "detail", args.detail_sheet)
        os.link(temporary, path)
        published = True
        temporary.unlink()
        fsync_directory(path.parent)
    except Exception:
        workbook.close()
        if temporary.exists():
            temporary.unlink()
        if published and path.exists():
            path.unlink()
            fsync_directory(path.parent)
        raise
    return {
        "ok": True,
        "workbook": str(path),
        "listing_sheet": args.listing_sheet,
        "detail_sheet": args.detail_sheet,
        "sha256": sha256_file(path),
    }


def unique_excel_target(directory: Path, base_name: str, profiles: dict[str, Any]) -> tuple[Path, str]:
    """Choose one suffix that is free for both the workbook and profile name."""
    raw_base = base_name[:-5] if base_name.casefold().endswith(".xlsx") else base_name
    raw_base = raw_base.strip()
    if not raw_base or Path(raw_base).name != raw_base:
        fail("Excel 기본 이름에는 폴더 경로를 넣을 수 없습니다", base_name=base_name)
    suffix = 1
    while True:
        profile_name = raw_base if suffix == 1 else f"{raw_base}-{suffix}"
        path = directory / f"{profile_name}.xlsx"
        if not path.exists() and profile_name not in profiles:
            return path, profile_name
        suffix += 1


def command_create_excel(args: argparse.Namespace, store_path: Path) -> dict[str, Any]:
    """Create, validate, register, and activate a new empty Excel ledger."""
    directory = resolve_path(args.directory)
    if not directory.is_dir():
        fail("저장할 기존 폴더를 찾을 수 없습니다", directory=str(directory))
    if not os.access(directory, os.W_OK):
        fail("저장할 폴더에 쓰기 권한이 없습니다", directory=str(directory))
    created_path: Path | None = None
    with file_lock(store_path):
        data = load_profile_store(store_path)
        previous_active = data.get("active_profile")
        path, profile_name = unique_excel_target(directory, args.base_name, data["profiles"])
        try:
            initialized = command_init_workbook(
                argparse.Namespace(
                    workbook=str(path),
                    listing_sheet=args.listing_sheet,
                    detail_sheet=args.detail_sheet,
                )
            )
            created_path = path
            validate_workbook_pair(path, args.listing_sheet, args.detail_sheet)
            profile = {
                "name": profile_name,
                "access": "local-xlsx",
                "label": profile_name,
                "updated_at": utc_now(),
                "workbook_path": str(path),
                "listing_sheet": args.listing_sheet,
                "detail_sheet": args.detail_sheet,
            }
            data["profiles"][profile_name] = profile
            data["active_profile"] = profile_name
            write_profile_store(store_path, data)
        except Exception:
            if created_path is not None and created_path.exists():
                created_path.unlink()
                fsync_directory(created_path.parent)
            raise
    return {
        **initialized,
        "profile_store": str(store_path),
        "profile": profile,
        "active_profile": profile_name,
        "previous_active_profile": previous_active,
    }


@contextmanager
def file_lock(target: Path) -> Iterable[None]:
    target.parent.mkdir(parents=True, exist_ok=True)
    lock_path = target.with_name(f".{target.name}.lock")
    with lock_path.open("a+", encoding="utf-8") as lock_handle:
        fcntl.flock(lock_handle.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(lock_handle.fileno(), fcntl.LOCK_UN)


def write_workbook_atomic(
    path: Path,
    sheet_name: str,
    columns: list[str],
    rows: list[dict[str, str]],
    kind: str,
    expected_sha: str,
    patches: list[dict[str, Any]],
) -> tuple[str, str]:
    """Apply only requested cell changes in a temporary workbook, then replace atomically."""
    require_openpyxl()
    require_xlsx_path(path)
    validate_rows(rows, kind)
    backup_dir = path.parent / ".maemul-backups"
    backup_dir.mkdir(mode=0o700, parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup = backup_dir / f"{path.name}.{timestamp}.bak"
    descriptor, raw_temporary = tempfile.mkstemp(prefix=f".{path.stem}.", suffix=".xlsx", dir=path.parent)
    os.close(descriptor)
    temporary = Path(raw_temporary)
    try:
        shutil.copy2(path, temporary)
        workbook = load_workbook(temporary)
        try:
            if sheet_name not in workbook.sheetnames:
                fail("Excel 매물장에 필요한 탭이 없습니다", path=str(path), sheet=sheet_name)
            sheet = workbook[sheet_name]
            physical_rows: dict[str, int] = {}
            for row_index in range(2, sheet.max_row + 1):
                item_id = workbook_cell_text(sheet.cell(row_index, 1).value).strip()
                if item_id:
                    physical_rows[item_id] = row_index
            column_indexes = {field: index for index, field in enumerate(columns, start=1)}
            for patch in patches:
                item_id = patch["item_id"]
                values = patch["values"]
                append = bool(patch.get("append"))
                if append:
                    if item_id in physical_rows:
                        fail("추가할 번호가 이미 Excel 탭에 있습니다", item_id=item_id)
                    row_index = max(physical_rows.values(), default=1) + 1
                    physical_rows[item_id] = row_index
                else:
                    row_index = physical_rows.get(item_id)
                    if row_index is None:
                        fail("수정할 번호가 Excel 탭에 없습니다", item_id=item_id)
                for field, value in values.items():
                    if field not in column_indexes:
                        fail("패치 필드가 스키마에 없습니다", field=field)
                    sheet.cell(row_index, column_indexes[field], value)
            if sheet.auto_filter.ref and physical_rows:
                min_col, min_row, max_col, max_row = range_boundaries(sheet.auto_filter.ref)
                max_row = max(max_row, max(physical_rows.values()))
                sheet.auto_filter.ref = (
                    f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                )
            workbook.save(temporary)
        finally:
            workbook.close()
        written_rows, _, _ = read_workbook_snapshot(temporary, kind, sheet_name)
        written_by_id = {row["번호"]: row for row in written_rows}
        for patch in patches:
            written = written_by_id.get(patch["item_id"])
            if written is None:
                fail("Excel 쓰기 검증에서 대상 행을 찾지 못했습니다", item_id=patch["item_id"])
            for field, expected in patch["values"].items():
                if written[field] != workbook_cell_text(expected):
                    fail(
                        "Excel 쓰기 검증 결과가 요청과 다릅니다",
                        item_id=patch["item_id"],
                        field=field,
                        expected=workbook_cell_text(expected),
                        actual=written[field],
                    )

        require_expected_hash(path, expected_sha)
        descriptor = os.open(backup, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
        with path.open("rb") as source, os.fdopen(descriptor, "wb") as destination:
            shutil.copyfileobj(source, destination)
            destination.flush()
            os.fsync(destination.fileno())
        if sha256_file(backup) != expected_sha:
            fail("백업 중 Excel 원본이 변경되었습니다", path=str(path), backup=str(backup))
        require_expected_hash(path, expected_sha)
        os.chmod(temporary, stat.S_IMODE(path.stat().st_mode))
        os.replace(temporary, path)
        fsync_directory(path.parent)
        fsync_directory(backup_dir)
    except Exception:
        if temporary.exists():
            temporary.unlink()
        raise
    return str(backup), sha256_file(path)


def empty_profile_store() -> dict[str, Any]:
    return {"version": 1, "active_profile": None, "profiles": {}}


def load_profile_store(path: Path) -> dict[str, Any]:
    if not path.exists():
        return empty_profile_store()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        fail("프로필 저장소를 읽을 수 없습니다", path=str(path), error=str(exc))
    if data.get("version") != 1 or not isinstance(data.get("profiles"), dict):
        fail("프로필 저장소 형식이 올바르지 않습니다", path=str(path))
    # 예전 버전의 '새 매물장 준비' 상태가 남아 있으면 직전 매물장으로 되돌린다.
    onboarding = data.pop("onboarding", None)
    if onboarding and not data.get("active_profile"):
        previous = onboarding.get("previous_active_profile")
        if previous in data["profiles"]:
            data["active_profile"] = previous
    validate_profile_store(data, path)
    return data


def write_profile_store(path: Path, data: dict[str, Any]) -> None:
    validate_profile_store(data, path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, raw_temp = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    temporary = Path(raw_temp)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(data, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.chmod(temporary, 0o600)
        os.replace(temporary, path)
        fsync_directory(path.parent)
    except Exception:
        if temporary.exists():
            temporary.unlink()
        raise


SHEET_ID_PATTERN = re.compile(r"^[A-Za-z0-9_-]+$")


def validate_profile_store(data: dict[str, Any], path: Path) -> None:
    if data.get("version") != 1 or not isinstance(data.get("profiles"), dict):
        fail("프로필 저장소 형식이 올바르지 않습니다", path=str(path))
    active = data.get("active_profile")
    if active is not None and (not isinstance(active, str) or active not in data["profiles"]):
        fail("활성 프로필이 저장소에 없습니다", path=str(path), name=active)
    for name, profile in data["profiles"].items():
        if not isinstance(name, str) or not name or not isinstance(profile, dict):
            fail("프로필 항목 형식이 올바르지 않습니다", path=str(path), name=name)
        access = profile.get("access")
        if access not in {"local-xlsx", "google-sheet"}:
            continue  # 구형 프로필은 show에서 지원 중단 안내를 제공한다.
        common = ("name", "access", "label", "updated_at", "listing_sheet", "detail_sheet")
        required = (*common, "workbook_path") if access == "local-xlsx" else (
            *common,
            "sheet_id",
            "spreadsheet_url",
            "connector",
            "account",
        )
        missing = [field for field in required if not isinstance(profile.get(field), str) or not profile[field].strip()]
        if missing:
            fail("프로필 필수 필드가 없습니다", path=str(path), name=name, fields=missing)
        if profile["name"] != name:
            fail("프로필 이름과 저장소 키가 다릅니다", path=str(path), name=name)
        if profile["listing_sheet"] == profile["detail_sheet"]:
            fail("프로필의 검색용 탭과 상세 탭 이름은 달라야 합니다", path=str(path), name=name)
        if access == "local-xlsx":
            workbook_path = Path(profile["workbook_path"])
            if not workbook_path.is_absolute() or workbook_path.suffix.casefold() != ".xlsx":
                fail("Excel 프로필 경로는 절대 .xlsx 경로여야 합니다", path=str(path), name=name)
        else:
            sheet_id = extract_sheet_id(profile["sheet_id"])
            if extract_sheet_id(profile["spreadsheet_url"]) != sheet_id:
                fail("Google 프로필의 시트 ID와 링크가 다릅니다", path=str(path), name=name)


def extract_sheet_id(raw: str) -> str:
    """Accept a bare Google Sheets ID or a full URL and return the ID."""
    value = (raw or "").strip()
    match = re.search(r"/spreadsheets/d/([A-Za-z0-9_-]+)", value)
    if match:
        return match.group(1)
    if not value or not SHEET_ID_PATTERN.match(value):
        fail("Google 시트 ID 형식이 올바르지 않습니다", value=raw, hint="시트 링크 또는 /d/와 /edit 사이의 ID를 전달하세요")
    return value


def profile_view(profile: dict[str, Any], verify_local: bool = False) -> dict[str, Any]:
    result = dict(profile)
    if profile.get("access") == "local-xlsx":
        workbook_path = Path(profile["workbook_path"])
        result["workbook_exists"] = workbook_path.is_file()
        if verify_local:
            validate_workbook_pair(workbook_path, profile["listing_sheet"], profile["detail_sheet"])
            result["workbook_valid"] = True
    return result


def command_profile(args: argparse.Namespace, store_path: Path) -> dict[str, Any]:
    if args.profile_action in {"list", "show"}:
        data = load_profile_store(store_path)
    else:
        with file_lock(store_path):
            data = load_profile_store(store_path)
            if args.profile_action == "set":
                if args.name in data["profiles"] and not args.replace:
                    fail("같은 이름의 프로필이 이미 있습니다", name=args.name, hint="확인 후 --replace를 사용하세요")
                profile: dict[str, Any] = {
                    "name": args.name,
                    "access": args.access,
                    "label": args.label or args.name,
                    "updated_at": utc_now(),
                }
                if args.access == "local-xlsx":
                    if not args.workbook:
                        fail("local-xlsx 프로필에는 --workbook이 필요합니다")
                    workbook_path = resolve_path(args.workbook)
                    validate_workbook_pair(workbook_path, args.listing_sheet, args.detail_sheet)
                    profile.update(
                        workbook_path=str(workbook_path),
                        listing_sheet=args.listing_sheet,
                        detail_sheet=args.detail_sheet,
                    )
                else:
                    if not args.sheet_id or not args.connector or not args.account:
                        fail("google-sheet 프로필에는 --sheet-id, --connector, --account가 필요합니다")
                    sheet_id = extract_sheet_id(args.sheet_id)
                    if args.spreadsheet_url:
                        url_id = extract_sheet_id(args.spreadsheet_url)
                        if url_id != sheet_id:
                            fail("시트 ID와 링크가 서로 다른 시트를 가리킵니다", sheet_id=sheet_id, url_sheet_id=url_id)
                    profile.update(
                        sheet_id=sheet_id,
                        spreadsheet_url=f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit",
                        connector=args.connector,
                        account=args.account,
                        listing_sheet=args.listing_sheet,
                        detail_sheet=args.detail_sheet,
                    )
                data["profiles"][args.name] = profile
                if args.activate or not data.get("active_profile"):
                    data["active_profile"] = args.name
                write_profile_store(store_path, data)
                return {
                    "ok": True,
                    "profile_store": str(store_path),
                    "profile": profile,
                    "active_profile": data["active_profile"],
                }
            if args.profile_action == "activate":
                if args.name not in data["profiles"]:
                    fail("프로필을 찾을 수 없습니다", name=args.name)
                verified = profile_view(data["profiles"][args.name], verify_local=True)
                data["active_profile"] = args.name
                write_profile_store(store_path, data)
                return {"ok": True, "active_profile": args.name, "profile": verified}

    if args.profile_action == "list":
        return {
            "ok": True,
            "profile_store": str(store_path),
            "active_profile": data.get("active_profile"),
            "profiles": [profile_view(profile) for profile in data["profiles"].values()],
        }

    name = args.name or data.get("active_profile")
    if not name:
        return {"ok": True, "profile_store": str(store_path), "active_profile": None, "profile": None}
    profile = data["profiles"].get(name)
    if not profile:
        fail("활성 프로필이 저장소에 없습니다", name=name)
    if profile.get("access") not in {"local-xlsx", "google-sheet"}:
        fail(
            "지원하지 않는 저장 방식의 프로필입니다",
            name=name,
            access=profile.get("access"),
            hint="Google Sheets 또는 로컬 Excel로 매물장을 다시 연결하세요",
        )
    return {"ok": True, "profile_store": str(store_path), "active_profile": name, "profile": profile_view(profile, verify_local=True)}


def compare_value(raw: str, criterion: dict[str, Any]) -> bool | None:
    if is_unknown(raw):
        return None
    field = criterion["field"]
    operator = criterion["op"]
    value = criterion.get("value")
    if operator not in SUPPORTED_OPERATORS:
        fail("지원하지 않는 검색 연산자입니다", operator=operator)

    if field in NUMERIC_FIELDS:
        left: Any = parse_number(raw, field)
        if operator == "between":
            if not isinstance(value, list) or len(value) != 2:
                fail("between 값은 두 원소 배열이어야 합니다", field=field)
            right: Any = [parse_number(value[0], field), parse_number(value[1], field)]
        elif operator in {"in", "not-in"}:
            if not isinstance(value, list):
                fail("in/not-in 값은 배열이어야 합니다", field=field)
            right = [parse_number(item, field) for item in value]
        else:
            right = parse_number(value, field)
    else:
        left = raw.strip().casefold()
        if operator == "between":
            if not isinstance(value, list) or len(value) != 2:
                fail("between 값은 두 원소 배열이어야 합니다", field=field)
            right = [str(value[0]).strip().casefold(), str(value[1]).strip().casefold()]
        elif operator in {"in", "not-in"}:
            if not isinstance(value, list):
                fail("in/not-in 값은 배열이어야 합니다", field=field)
            right = [str(item).strip().casefold() for item in value]
        else:
            right = str(value).strip().casefold()

    if operator == "eq":
        return left == right
    if operator == "ne":
        return left != right
    if operator == "in":
        return left in right
    if operator == "not-in":
        return left not in right
    if operator == "contains":
        return right in left
    if operator == "lte":
        return left <= right
    if operator == "gte":
        return left >= right
    return right[0] <= left <= right[1]


def validate_criteria(criteria: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not isinstance(criteria, dict):
        fail("검색 조건은 JSON 객체여야 합니다")
    hard = criteria.get("hard", [])
    soft = criteria.get("soft", [])
    if not isinstance(hard, list) or not isinstance(soft, list):
        fail("hard와 soft는 배열이어야 합니다")
    for criterion in [*hard, *soft]:
        if not isinstance(criterion, dict):
            fail("각 검색 조건은 객체여야 합니다")
        if criterion.get("field") not in LISTING_COLUMNS:
            fail("검색 조건 필드가 스키마에 없습니다", field=criterion.get("field"))
        if criterion.get("op") not in SUPPORTED_OPERATORS:
            fail("지원하지 않는 검색 연산자입니다", operator=criterion.get("op"))
        if "value" not in criterion:
            fail("검색 조건에 value가 없습니다", criterion=criterion)
        if criterion.get("value") is None:
            fail("검색 조건 value는 null일 수 없습니다", criterion=criterion)
        if criterion["field"] in NUMERIC_FIELDS and criterion["op"] == "contains":
            fail("숫자 필드에는 contains를 사용할 수 없습니다", field=criterion["field"])
        if criterion["field"] not in NUMERIC_FIELDS and criterion["op"] in ORDERED_ONLY_OPERATORS:
            fail(
                "텍스트 필드에는 크기 비교 연산자를 사용할 수 없습니다",
                field=criterion["field"],
                operator=criterion["op"],
                hint="eq·ne·in·not-in·contains 중에서 고르세요",
            )
    return hard, soft


def classify_row(row: dict[str, str], hard: list[dict[str, Any]]) -> tuple[str, list[str], list[str]]:
    unknown: list[str] = []
    failed: list[str] = []
    for criterion in hard:
        result = compare_value(row[criterion["field"]], criterion)
        label = criterion.get("label") or criterion["field"]
        if result is None:
            unknown.append(label)
        elif not result:
            failed.append(label)
    if failed:
        return "excluded", unknown, failed
    if unknown:
        return "needs_verification", unknown, failed
    return "match", unknown, failed


def transaction_price(row: dict[str, str]) -> tuple[float, float]:
    infinity = float("inf")
    transaction = row["거래"]
    if transaction == "매매":
        return (infinity if is_unknown(row["매매가(만원)"]) else float(row["매매가(만원)"]), 0.0)
    if transaction == "전세":
        return (infinity if is_unknown(row["보증금(만원)"]) else float(row["보증금(만원)"]), 0.0)
    deposit = infinity if is_unknown(row["보증금(만원)"]) else float(row["보증금(만원)"])
    rent = infinity if is_unknown(row["월세(만원)"]) else float(row["월세(만원)"])
    # 월세 손님이 먼저 보는 값은 월세다. 보증금은 협상 여지가 큰 편이므로 2차 키로 둔다.
    return rent, deposit


def receipt_rank(row: dict[str, str]) -> int:
    raw = row["접수일"]
    if is_unknown(raw):
        return 0
    try:
        return int(raw.replace("-", ""))
    except ValueError:
        return 0


def decorated_row(row: dict[str, str], soft: list[dict[str, Any]], unknown: list[str]) -> dict[str, Any]:
    satisfied: list[str] = []
    soft_unknown: list[str] = []
    for criterion in soft:
        result = compare_value(row[criterion["field"]], criterion)
        label = criterion.get("label") or criterion["field"]
        if result is True:
            satisfied.append(label)
        elif result is None:
            soft_unknown.append(label)
    return {
        "row": row,
        "soft_score": len(satisfied),
        "soft_satisfied": satisfied,
        "soft_unknown": soft_unknown,
        "hard_unknown": unknown,
    }


TRANSACTION_ORDER = {"매매": 0, "전세": 1, "월세": 2}


def transaction_group(row: dict[str, str]) -> tuple[int, str]:
    """거래유형별 정렬 그룹. 스키마 밖 값은 뒤로 보내되 순서를 결정적으로 유지한다."""
    transaction = row["거래"]
    return (TRANSACTION_ORDER.get(transaction, len(TRANSACTION_ORDER)), transaction)


def search_rows(
    rows: list[dict[str, str]],
    hard: list[dict[str, Any]],
    soft: list[dict[str, Any]],
    include_hold: bool,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    matches: list[dict[str, Any]] = []
    possible: list[dict[str, Any]] = []
    for row in rows:
        if row["상태"] == "완료" or (row["상태"] == "보류" and not include_hold):
            continue
        classification, unknown, _ = classify_row(row, hard)
        if classification == "excluded":
            continue
        decorated = decorated_row(row, soft, unknown)
        if classification == "match":
            matches.append(decorated)
        else:
            possible.append(decorated)

    found_transactions = {
        item["row"]["거래"]
        for item in [*matches, *possible]
        if not is_unknown(item["row"]["거래"])
    }
    group_by_transaction = len(found_transactions) > 1

    def sort_key(item: dict[str, Any]) -> tuple[Any, ...]:
        row = item["row"]
        price_a, price_b = transaction_price(row)
        group = transaction_group(row) if group_by_transaction else (0, "")
        return (group, -item["soft_score"], price_a, price_b, -receipt_rank(row), row["번호"])

    matches.sort(key=sort_key)
    possible.sort(key=sort_key)
    return matches, possible


def relaxation_counts(
    rows: list[dict[str, str]],
    hard: list[dict[str, Any]],
    include_hold: bool,
) -> list[dict[str, Any]]:
    """Count all one-condition relaxations in one pass over the rows."""
    counts = [[0, 0] for _ in hard]
    for row in rows:
        if row["상태"] == "완료" or (row["상태"] == "보류" and not include_hold):
            continue
        results = [compare_value(row[criterion["field"]], criterion) for criterion in hard]
        for removed_index in range(len(hard)):
            remaining = results[:removed_index] + results[removed_index + 1 :]
            if False in remaining:
                continue
            counts[removed_index][1 if None in remaining else 0] += 1
    return [
        {
            "removed": criterion.get("label") or criterion["field"],
            "verified_count": counts[index][0],
            "needs_verification_count": counts[index][1],
        }
        for index, criterion in enumerate(hard)
    ]


def read_command_table(args: argparse.Namespace, kind: str) -> tuple[Path, str | None, list[dict[str, str]], dict[str, Any], str]:
    workbook_value = getattr(args, "workbook", None)
    if workbook_value:
        path = resolve_path(workbook_value)
        sheet_name = getattr(args, "sheet", None) or ("매물" if kind == "listing" else "매물상세")
        rows, summary, snapshot_sha = read_workbook_snapshot(path, kind, sheet_name)
        return path, sheet_name, rows, summary, snapshot_sha
    path = resolve_path(args.file)
    rows, summary, snapshot_sha = read_csv_snapshot(path, kind)
    return path, None, rows, summary, snapshot_sha


def target_metadata(path: Path, sheet_name: str | None) -> dict[str, Any]:
    if sheet_name:
        return {"workbook": str(path), "sheet": sheet_name}
    return {"file": str(path)}


def command_search(args: argparse.Namespace) -> dict[str, Any]:
    path, sheet_name, rows, summary, snapshot_sha = read_command_table(args, "listing")
    criteria = parse_json(args.criteria_json, "검색 조건")
    hard, soft = validate_criteria(criteria)
    matches, possible = search_rows(rows, hard, soft, args.include_hold)

    relaxations = relaxation_counts(rows, hard, args.include_hold)

    warnings: list[dict[str, Any]] = []
    found = sorted(
        {item["row"]["거래"] for item in [*matches, *possible] if not is_unknown(item["row"]["거래"])},
        key=lambda transaction: (TRANSACTION_ORDER.get(transaction, len(TRANSACTION_ORDER)), transaction),
    )
    if len(found) > 1:
        warnings.append(
            {
                "warning": "거래유형이 섞여 있어 유형 간 가격 순위는 비교할 수 없습니다",
                "transactions": found,
                "hint": "한 거래유형만 남도록 조건을 좁히면 유형 안에서만 가격 순위를 매깁니다",
            }
        )

    match_status_counts = {
        status: sum(1 for item in matches if item["row"]["상태"] == status)
        for status in ("진행", "보류")
    }
    verification_status_counts = {
        status: sum(1 for item in possible if item["row"]["상태"] == status)
        for status in ("진행", "보류")
    }

    return {
        "ok": True,
        **target_metadata(path, sheet_name),
        "sha256": snapshot_sha,
        "validated_rows": summary["rows"],
        "matches_count": len(matches),
        "needs_verification_count": len(possible),
        "match_status_counts": match_status_counts,
        "verification_status_counts": verification_status_counts,
        "matches": matches[: args.limit],
        "needs_verification": possible[: args.limit],
        "relaxations": relaxations,
        "warnings": warnings,
    }


def require_expected_hash(path: Path, expected: str) -> None:
    actual = sha256_file(path)
    if actual != expected:
        fail("파일이 확인 이후 변경되었습니다", path=str(path), expected_sha=expected, actual_sha=actual)


def find_row(rows: list[dict[str, str]], item_id: str) -> tuple[int, dict[str, str]]:
    for index, row in enumerate(rows):
        if row["번호"] == item_id:
            return index, row
    fail("번호에 해당하는 행을 찾을 수 없습니다", item_id=item_id)


def duplicate_signatures(row: dict[str, str]) -> list[tuple[str, tuple[str, ...]]]:
    """같은 물건이 다른 번호로 다시 들어오는 경우를 찾기 위한 지문.

    여러 출처(내 장부·단톡방·플랫폼)에서 같은 매물이 들어오는 것이 이 스킬의 정상 입력이므로
    저장을 막지 않고 후보만 알린다. 확정 판정이 아니다.
    """
    def normalized(field: str, value: str) -> str:
        text = re.sub(r"\s+", "", str(value)).casefold()
        if field == "동호":
            parts = re.findall(r"\d+(?:\.0+)?", text)
            if len(parts) == 2:
                return "-".join(format(float(part), ".15g") for part in parts)
        if field in NUMERIC_FIELDS and not is_unknown(text):
            try:
                number = float(text)
            except ValueError:
                return text
            if math.isfinite(number):
                return format(number, ".15g")
        return text

    signatures: list[tuple[str, tuple[str, ...]]] = []
    complex_name, unit = row["단지명"], row["동호"]
    if not is_unknown(row["지역"]) and not is_unknown(complex_name) and not is_unknown(unit):
        signatures.append(
            (
                "지역·단지·동호",
                tuple(normalized(field, row[field]) for field in ("지역", "단지명", "동호")),
            )
        )
    core_fields = ("지역", "동네", "거래", "전용(㎡)", "매매가(만원)", "보증금(만원)", "월세(만원)")
    if all(not is_unknown(value) for value in (row["지역"], row["거래"], row["전용(㎡)"])):
        signatures.append(("지역·거래·면적·가격", tuple(normalized(field, row[field]) for field in core_fields)))
    return signatures


def duplicate_candidates(row: dict[str, str], rows: list[dict[str, str]]) -> list[dict[str, str]]:
    incoming = duplicate_signatures(row)
    if not incoming:
        return []
    hits: list[dict[str, str]] = []
    for existing in rows:
        if existing["상태"] == "완료":
            continue
        existing_signatures = duplicate_signatures(existing)
        for label, signature in incoming:
            if any(label == other_label and signature == other for other_label, other in existing_signatures):
                hits.append({"item_id": existing["번호"], "matched_on": label})
                break
    return hits


def generated_id(rows: list[dict[str, str]]) -> str:
    numbers: list[int] = []
    for row in rows:
        item_id = row["번호"]
        if item_id.startswith("P") and item_id[1:].isdigit():
            numbers.append(int(item_id[1:]))
    return f"P{(max(numbers, default=0) + 1):03d}"


def prepare_listing_additions(
    rows: list[dict[str, str]], payload: Any
) -> tuple[list[dict[str, str]], list[dict[str, Any]], list[dict[str, Any]]]:
    records = payload if isinstance(payload, list) else [payload]
    if not records or len(records) > 1000 or any(not isinstance(record, dict) for record in records):
        fail("매물 JSON은 객체 또는 1~1000개의 객체 배열이어야 합니다")
    added: list[dict[str, str]] = []
    warnings: list[dict[str, Any]] = []
    patches: list[dict[str, Any]] = []
    for record in records:
        unknown_fields = sorted(set(record) - set(LISTING_COLUMNS))
        if unknown_fields:
            fail("스키마에 없는 컬럼입니다", fields=unknown_fields)
        row = {column: "?" for column in LISTING_COLUMNS}
        for field in ("매매가(만원)", "보증금(만원)", "월세(만원)"):
            row[field] = ""
        row.update({key: normalize_input_value(key, value) for key, value in record.items()})
        row["번호"] = row["번호"] if not is_unknown(row["번호"]) else generated_id(rows)
        if "상태" in record and row["상태"] not in STATUS_VALUES:
            fail("명시한 상태 값이 올바르지 않습니다", value=row["상태"], allowed=sorted(STATUS_VALUES))
        missing_core = [field for field in ("종류", "거래", "지역") if is_unknown(row[field])]
        if missing_core:
            row["상태"] = "보류"
            warnings.append({"item_id": row["번호"], "fields": missing_core, "warning": "검색 핵심값 미확인으로 보류"})
        elif "상태" not in record:
            row["상태"] = "진행"
        if "접수일" not in record:
            row["접수일"] = date.today().isoformat()
        if any(existing["번호"] == row["번호"] for existing in rows):
            fail("이미 존재하는 번호입니다", item_id=row["번호"])
        similar = duplicate_candidates(row, rows)
        if similar:
            warnings.append(
                {
                    "item_id": row["번호"],
                    "warning": "같은 물건일 수 있는 기존 매물이 있습니다",
                    "candidates": similar,
                }
            )
        rows.append(row)
        added.append(row)
        patches.append({"item_id": row["번호"], "values": dict(row), "append": True})
    validate_rows(rows, "listing")
    return added, warnings, patches


def parse_existing_rows_json(raw: str) -> list[dict[str, str]]:
    payload = parse_json(raw, "기존 매물")
    if not isinstance(payload, list) or any(not isinstance(row, dict) for row in payload):
        fail("기존 매물 JSON은 객체 배열이어야 합니다")
    rows = [
        {column: workbook_cell_text(row.get(column, "")) for column in LISTING_COLUMNS}
        for row in payload
    ]
    validate_rows(rows, "listing")
    return rows


def command_prepare_add(args: argparse.Namespace) -> dict[str, Any]:
    rows = parse_existing_rows_json(args.existing_json)
    payload = parse_json(args.record_json, "매물")
    added, warnings, _ = prepare_listing_additions(rows, payload)
    return {"ok": True, "prepared_count": len(added), "prepared": added, "warnings": warnings}


def command_mutation(args: argparse.Namespace) -> dict[str, Any]:
    path = resolve_path(args.workbook)
    with file_lock(path):
        path, sheet_name, rows, _, snapshot_sha = read_command_table(args, "listing")
        if snapshot_sha != args.expected_sha:
            fail(
                "파일이 확인 이후 변경되었습니다",
                path=str(path),
                expected_sha=args.expected_sha,
                actual_sha=snapshot_sha,
            )
        warnings: list[dict[str, Any]] = []
        patches: list[dict[str, Any]] = []
        if args.command == "add":
            payload = parse_json(args.record_json, "매물")
            added, warnings, patches = prepare_listing_additions(rows, payload)
            changed = added[0] if isinstance(payload, dict) else added
        else:
            index, current = find_row(rows, args.id)
            before = dict(current)
            if args.command == "complete":
                if current["상태"] == "완료":
                    fail("이미 완료 상태입니다", item_id=args.id)
                current["상태"] = "완료"
                patch_values = {"상태": "완료"}
            else:
                changes = parse_json(args.changes_json, "변경값")
                if not isinstance(changes, dict):
                    fail("변경값 JSON은 객체여야 합니다")
                unknown_fields = sorted(set(changes) - set(LISTING_COLUMNS))
                if unknown_fields:
                    fail("스키마에 없는 컬럼입니다", fields=unknown_fields)
                if "번호" in changes:
                    fail("번호는 update로 변경할 수 없습니다")
                patch_values = {key: normalize_input_value(key, value) for key, value in changes.items()}
                if "상태" in patch_values and patch_values["상태"] not in STATUS_VALUES:
                    fail("명시한 상태 값이 올바르지 않습니다", value=patch_values["상태"], allowed=sorted(STATUS_VALUES))
                current.update(patch_values)
                missing_core = [field for field in ("종류", "거래", "지역") if is_unknown(current[field])]
                if missing_core:
                    current["상태"] = "보류"
                    patch_values["상태"] = "보류"
                    warnings.append({"item_id": current["번호"], "fields": missing_core, "warning": "검색 핵심값 미확인으로 보류"})
            rows[index] = current
            changed = {"before": before, "after": current}
            patches.append({"item_id": current["번호"], "values": patch_values, "append": False})

        backup, new_hash = write_workbook_atomic(
            path, sheet_name, LISTING_COLUMNS, rows, "listing", args.expected_sha, patches
        )
        return {
            "ok": True,
            **target_metadata(path, sheet_name),
            "backup": backup,
            "sha256": new_hash,
            "changed": changed,
            "warnings": warnings,
        }


def command_detail_upsert(args: argparse.Namespace) -> dict[str, Any]:
    path = resolve_path(args.workbook)
    with file_lock(path):
        path, sheet_name, rows, _, snapshot_sha = read_command_table(args, "detail")
        if snapshot_sha != args.expected_sha:
            fail(
                "파일이 확인 이후 변경되었습니다",
                path=str(path),
                expected_sha=args.expected_sha,
                actual_sha=snapshot_sha,
            )
        listing_rows, _, listing_sha = read_workbook_snapshot(path, "listing", args.listing_sheet)
        if listing_sha != snapshot_sha:
            fail("상세 확인 중 Excel 파일이 변경되었습니다", path=str(path))
        if not any(row["번호"] == args.id for row in listing_rows):
            fail("매물 탭에 없는 번호의 상세행은 만들 수 없습니다", item_id=args.id, listing_sheet=args.listing_sheet)
        changes = parse_json(args.changes_json, "상세 변경값")
        if not isinstance(changes, dict):
            fail("상세 변경값 JSON은 객체여야 합니다")
        unknown_fields = sorted(set(changes) - set(DETAIL_COLUMNS))
        if unknown_fields:
            fail("상세 스키마에 없는 컬럼입니다", fields=unknown_fields)
        if "번호" in changes and normalize_input_value("번호", changes["번호"]) != args.id:
            fail("상세 번호는 대상 번호와 같아야 합니다")
        try:
            index, current = find_row(rows, args.id)
            before: dict[str, str] | None = dict(current)
        except ToolError:
            index = len(rows)
            current = {column: "?" for column in DETAIL_COLUMNS}
            current["번호"] = args.id
            before = None
        current.update({key: normalize_input_value(key, value) for key, value in changes.items()})
        current["번호"] = args.id
        append = index == len(rows)
        if index == len(rows):
            rows.append(current)
        else:
            rows[index] = current
        patch_values = {key: normalize_input_value(key, value) for key, value in changes.items()}
        if append:
            patch_values = dict(current)
        backup, new_hash = write_workbook_atomic(
            path,
            sheet_name,
            DETAIL_COLUMNS,
            rows,
            "detail",
            args.expected_sha,
            [{"item_id": args.id, "values": patch_values, "append": append}],
        )
        return {
            "ok": True,
            **target_metadata(path, sheet_name),
            "backup": backup,
            "sha256": new_hash,
            "changed": {"before": before, "after": current},
        }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--profile-store", default=None, help="Override profile store path for tests or isolated use")
    subparsers = parser.add_subparsers(dest="command", required=True)

    def add_table_target(command_parser: argparse.ArgumentParser) -> None:
        """Read-only commands accept a local .xlsx workbook or a standard CSV snapshot."""
        target = command_parser.add_mutually_exclusive_group(required=True)
        target.add_argument("--file", help="Standard CSV snapshot path (read-only use)")
        target.add_argument("--workbook", help="Local .xlsx workbook path")
        command_parser.add_argument("--sheet", help="Workbook sheet name; defaults to 매물 or 매물상세")

    def add_workbook_target(command_parser: argparse.ArgumentParser) -> None:
        """Mutating commands only write to a local .xlsx workbook."""
        command_parser.add_argument("--workbook", required=True, help="Local .xlsx workbook path")
        command_parser.add_argument("--sheet", help="Workbook sheet name; defaults to 매물 or 매물상세")

    profile = subparsers.add_parser("profile")
    profile_actions = profile.add_subparsers(dest="profile_action", required=True)
    profile_actions.add_parser("list")
    profile_show = profile_actions.add_parser("show")
    profile_show.add_argument("--name")
    profile_activate = profile_actions.add_parser("activate")
    profile_activate.add_argument("--name", required=True)
    profile_set = profile_actions.add_parser("set")
    profile_set.add_argument("--name", required=True)
    profile_set.add_argument("--access", choices=["local-xlsx", "google-sheet"], required=True)
    profile_set.add_argument("--workbook")
    profile_set.add_argument("--sheet-id")
    profile_set.add_argument("--spreadsheet-url")
    profile_set.add_argument("--connector")
    profile_set.add_argument("--account")
    profile_set.add_argument("--listing-sheet", default="매물")
    profile_set.add_argument("--detail-sheet", default="매물상세")
    profile_set.add_argument("--label")
    profile_set.add_argument("--activate", action="store_true")
    profile_set.add_argument("--replace", action="store_true")

    init_workbook = subparsers.add_parser("init-workbook")
    init_workbook.add_argument("--workbook", required=True)
    init_workbook.add_argument("--listing-sheet", default="매물")
    init_workbook.add_argument("--detail-sheet", default="매물상세")

    create_excel = subparsers.add_parser("create-excel")
    create_excel.add_argument("--directory", required=True)
    create_excel.add_argument("--base-name", default="매물장")
    create_excel.add_argument("--listing-sheet", default="매물")
    create_excel.add_argument("--detail-sheet", default="매물상세")

    validate = subparsers.add_parser("validate")
    add_table_target(validate)
    validate.add_argument("--kind", choices=["listing", "detail"], required=True)

    file_hash = subparsers.add_parser("hash")
    file_hash.add_argument("--file", required=True)

    inspect = subparsers.add_parser("inspect")
    add_table_target(inspect)
    inspect.add_argument("--id", required=True)

    search = subparsers.add_parser("search")
    add_table_target(search)
    search.add_argument("--criteria-json", required=True)
    search.add_argument("--limit", type=int, default=10)
    search.add_argument("--include-hold", action="store_true")

    prepare_add = subparsers.add_parser("prepare-add")
    prepare_add.add_argument("--existing-json", required=True)
    prepare_add.add_argument("--record-json", required=True)

    add = subparsers.add_parser("add")
    add_workbook_target(add)
    add.add_argument("--record-json", required=True)
    add.add_argument("--expected-sha", required=True)

    update = subparsers.add_parser("update")
    add_workbook_target(update)
    update.add_argument("--id", required=True)
    update.add_argument("--changes-json", required=True)
    update.add_argument("--expected-sha", required=True)

    complete = subparsers.add_parser("complete")
    add_workbook_target(complete)
    complete.add_argument("--id", required=True)
    complete.add_argument("--expected-sha", required=True)

    detail = subparsers.add_parser("detail-upsert")
    add_workbook_target(detail)
    detail.add_argument("--id", required=True)
    detail.add_argument("--changes-json", required=True)
    detail.add_argument("--expected-sha", required=True)
    detail.add_argument("--listing-sheet", default="매물")
    return parser


def run(args: argparse.Namespace) -> dict[str, Any]:
    store_path = resolve_path(args.profile_store) if args.profile_store else default_profile_store()
    if args.command == "profile":
        return command_profile(args, store_path)
    if args.command == "init-workbook":
        return command_init_workbook(args)
    if args.command == "create-excel":
        return command_create_excel(args, store_path)
    if args.command == "validate":
        path, sheet_name, _, summary, snapshot_sha = read_command_table(args, args.kind)
        return {"ok": True, **target_metadata(path, sheet_name), "kind": args.kind, "sha256": snapshot_sha, **summary}
    if args.command == "hash":
        path = resolve_path(args.file)
        if not path.is_file():
            fail("파일을 찾을 수 없습니다", path=str(path))
        return {"ok": True, "file": str(path), "sha256": sha256_file(path)}
    if args.command == "inspect":
        path, sheet_name, rows, _, snapshot_sha = read_command_table(args, "listing")
        _, row = find_row(rows, args.id)
        return {"ok": True, **target_metadata(path, sheet_name), "sha256": snapshot_sha, "row": row}
    if args.command == "search":
        if args.limit < 1 or args.limit > 100:
            fail("limit은 1부터 100 사이여야 합니다")
        return command_search(args)
    if args.command == "prepare-add":
        return command_prepare_add(args)
    if args.command in {"add", "update", "complete"}:
        return command_mutation(args)
    if args.command == "detail-upsert":
        return command_detail_upsert(args)
    fail("알 수 없는 명령입니다", command=args.command)


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        emit(run(args))
        return 0
    except ToolError as exc:
        try:
            details = json.loads(str(exc))
        except json.JSONDecodeError:
            details = {"message": str(exc)}
        emit({"ok": False, "error": details})
        return 2
    except Exception as exc:  # pragma: no cover - defensive CLI boundary
        emit({"ok": False, "error": {"message": "예상하지 못한 오류", "type": type(exc).__name__, "detail": str(exc)}})
        return 3


if __name__ == "__main__":
    sys.exit(main())
